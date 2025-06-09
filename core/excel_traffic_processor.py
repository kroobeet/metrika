"""Модуль для обработки и агрегирования данных о трафике из JSON в Excel."""

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, TypedDict

from openpyxl import load_workbook


class TrafficData(TypedDict):
    """Определение типа данных источника трафика."""
    organic: int
    direct: int
    ad: int
    internal: int
    referral: int
    recommendation: int
    social: int


@dataclass
class ExcelTrafficConfig:
    """Конфигурация для ExcelTrafficProcessor."""
    json_path: Path = Path("temp/api_response.json")
    excel_file_path: Path = Path("")
    sheet_name: str = "Сводка"
    city_column: str = "B"
    traffic_columns: Dict[str, str] = None


class ExcelTrafficProcessor:
    """Обрабатывает и агрегирует данные о трафике из JSON в Excel"""

    DEFAULT_TRAFFIC_COLUMNS = {
        "organic": "G",         # Поисковые системы
        "direct": "H",          # Прямые переходы
        "ad": "I",              # Реклама
        "internal": "J",        # Внутренние переходы
        "referral": "K",        # Переходы по ссылкам
        "recommendation": "L",  # Рекомендации
        "social": "M",          # Социальные сети
    }

    def __init__(self, config: Optional[ExcelTrafficConfig] = None):
        """Инициализация с конфигурацией"""
        self.config = config or ExcelTrafficConfig()
        self.traffic_data: Dict[str, TrafficData] = {}

        if self.config.traffic_columns is None:
            self.config.traffic_columns = self.DEFAULT_TRAFFIC_COLUMNS

    @staticmethod
    def extract_city_name(full_name: str) -> str:
        """Извлекает название города из строки регион-город.

        Args:
            full_name: Строка в формате "Регион - Город"

        Returns:
            Извлеченное название города или исходная строка, если разделитель не найден
        """
        return full_name.split(" - ",)[-1]

    def load_traffic_data(self) -> None:
        """Загрузка и агрегация данных о трафике из файла JSON"""
        with open(self.config.json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Инициализация структуры данных
        self.traffic_data = {
            self.extract_city_name(region_city): {
                key: 0 for key in self.config.traffic_columns
            }
            for region_city in data
        }

        # Агрегация данных
        for region_city, region_data in data.items():
            city = self.extract_city_name(region_city)
            for item in region_data["data"]:
                if len(item["dimensions"]) > 1:
                    traffic_type = item["dimensions"][1]["id"]
                    visits = item["metrics"][0]
                    if traffic_type in self.traffic_data[city]:
                        self.traffic_data[city][traffic_type] += visits

    def update_excel(self, output_path: Path) -> None:
        """Обновляем файл Excel с данными о трафике.

        Args:
            output_path: Путь для сохранения обновлённого файла Excel.
        """
        wb = load_workbook(self.config.excel_file_path)
        sheet = wb[self.config.sheet_name]

        for row in range(1, sheet.max_row + 1):
            city = sheet[f"{self.config.city_column}{row}"].value
            if city in self.traffic_data:
                for traffic_type, column in self.config.traffic_columns.items():
                    sheet[f"{column}{row}"] = self.traffic_data[city][traffic_type]

        wb.save(output_path)

    def process(self, output_path: Path = Path("metrika_report.xlsx")) -> None:
        """Запуск технологического процесса"""
        self.load_traffic_data()
        self.update_excel(output_path)
