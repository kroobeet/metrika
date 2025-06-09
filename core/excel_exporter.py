"""Модуль для экспорта отчёта по полученным данным в Excel."""

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook

from .data_processor import DataProcessor
from .models import ReportData
from .exceptions import ExcelExportError
from .excel_traffic_processor import ExcelTrafficProcessor, ExcelTrafficConfig

logger = logging.getLogger(__name__)


@dataclass
class ExcelExportConfig:
    """Конфигурация для экспортера Excel."""
    header_font: Font = Font(bold=True)
    center_alignment: Alignment = Alignment(horizontal='center')
    sheet_name_max_length: int = 31
    temp_dir: Path = Path("temp")
    

class ExcelExporter:
    """Управляет экспортом отчёта по данным в формат Excel."""
    
    TRAFFIC_HEADERS = {
        'organic': "Поисковые системы",
        'direct': "Прямые переходы",
        'ad': "Реклама",
        'internal': "Внутренние переходы",
        'referral': "Переходы по ссылкам",
        'recommendation': "Рекомендации",
        'social': "Социальные сети"
    }
    
    def __init__(self, config: Optional[ExcelExportConfig] = None):
        """Инициализация с конфигурацией"""
        self.config = config or ExcelExportConfig()
        self._ensure_temp_dir_exists()
        
    def _ensure_temp_dir_exists(self) -> None:
        """Убедимся, что временный каталог существует."""
        self.config.temp_dir.mkdir(parents=True, exist_ok=True)
    
    def export_report(
            self,
            data: Dict[str, List[ReportData]],
            file_path: Path,
            date_from: str,
            date_to: str,
            filters: str = ""
    ) -> None:
        """Экспортирует отчёт по данным в файл Excel.
        
        Args:
            data: Словарь с названиями локаций и данными из отчёта
            file_path: Путь для сохранения файла Excel
            date_from: Начальная дата периода отчёта
            date_to: Конечная дата периода отчёта
            filters: Описание применяемых фильтров
            
        Raises:
            ExcelExportError: Если экспорт завершится неудачей
        """
        temp_file = self.config.temp_dir / f"temp_{file_path.name}"
        
        try:
            # Create initial workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            self._create_summary_sheet(wb, data, date_from, date_to, filters)
            
            for location, report_data in data.items():
                self._create_city_sheet(wb, location, report_data, date_from, date_to, filters)
                
            wb.save(temp_file)
            logger.info("Временный отчёт экспортирован в %s", temp_file)
            
            # Process traffic data
            traffic_processor = ExcelTrafficProcessor(ExcelTrafficConfig(excel_file_path=temp_file))
            traffic_processor.process(output_path=file_path)
            
            temp_file.unlink()
            logger.info("Временный файл %s удалён", temp_file)
            
        except Exception as e:
            logger.error("Ошибка экспорта: %s", str(e), exc_info=True)
            raise ExcelExportError(f"Экспорт не выполнен: {str(e)}") from e
        
    def _create_summary_sheet(
            self,
            wb: Workbook,
            data: Dict[str, List[ReportData]],
            date_from: str,
            date_to: str,
            filters: str
    ) -> None:
        """Создание листа со сводной в файле Excel"""
        ws = wb.create_sheet(title="Сводка")
        
        # Добавляем заголовки отчёта
        ws.append([f"Отчет за период с {date_from} по {date_to}"])
        ws.append(["Отчет: Яндекс.Метрика - Аналитика по городам"])

        if filters:
            ws.append([f"Фильтры: {filters}"])

        ws.append([])

        # Добавляем заголовки таблицы
        headers = [
            "Регион", "Город", "Посещения", "Посетители", "Просмотры страниц", "Глубина просмотра"
        ]
        headers.extend(self.TRAFFIC_HEADERS.values())
        ws.append(headers)

        # Стили заголовков
        for cell in ws[ws.max_row]:
            cell.font = self.config.header_font
            cell.alignment = self.config.center_alignment

        # Добавление строк с данными
        for location, report_data in data.items():
            region, city = location.split(" - ")
            totals = DataProcessor.calculate_totals(report_data)

            row_data = [
                region,
                city,
                totals['all']['visits'],
                totals['all']['users'],
                totals['all']['pageviews'],
                totals['all']['pageviews'] / totals['all']['visits'] if totals['all']['visits'] > 0 else 0
            ]

            # Добавление данных по источникам трафика
            row_data.extend(
                totals['sources'].get(source, {}).get('visits', 0)
                for source in self.TRAFFIC_HEADERS
            )

            ws.append(row_data)

    def _create_city_sheet(
            self,
            wb: Workbook,
            location: str,
            report_data: List[ReportData],
            date_from: str,
            date_to: str,
            filters: str
    ) -> None:
        """Создание листа для каждого города"""
        _, city = location.split(" - ")
        sheet_name = city[:self.config.sheet_name_max_length]
        ws = wb.create_sheet(title=sheet_name)

        # Добавляем заголовки отчёта
        ws.append([f"Отчет за период с {date_from} по {date_to}"])
        ws.append([f"Отчет: Яндекс.Метрика - {city}"])

        if filters:
            ws.append([f"Фильтры: {filters} и город = '{city}'"])

        ws.append([])

        # Добавляем заголовки таблиц
        headers = ["Дата", "Источник трафика", "Посещения", "Посетители", "Глубина просмотра"]
        ws.append(headers)

        # Стили заголовков
        for cell in ws[ws.max_row]:
            cell.font = self.config.header_font
            cell.alignment = self.config.center_alignment

        # Добавляем строки с данными
        for item in report_data:
            source = self.TRAFFIC_HEADERS.get(
                item.traffic_source,
                item.traffic_source or "Other"
            )
            ws.append([
                item.date.strftime("%Y-%m-%d"),
                source,
                item.visits,
                item.users,
                item.pageviews / item.visits if item.visits > 0 else 0,
            ])
